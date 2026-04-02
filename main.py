# ============================================================
# SISTEMA DE REGISTRO DIARIO v2 - BACKEND FASTAPI
# main.py
# pip install fastapi uvicorn sqlalchemy psycopg2-binary
#             python-jose passlib[bcrypt] python-dotenv openpyxl
# uvicorn main:app --reload --host 0.0.0.0 --port 8000
# ============================================================

from fastapi import FastAPI, Depends, HTTPException, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from sqlalchemy import create_engine, Column, Integer, String, Numeric, Boolean, Date, Time, DateTime, ForeignKey, func, text
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session, relationship
from passlib.context import CryptContext
from jose import JWTError, jwt
from datetime import date, datetime, timedelta
from typing import Optional, List
from pydantic import BaseModel
import os, io
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://postgres:postgres@db:5432/registro_diario")
SECRET_KEY   = os.getenv("SECRET_KEY", "cambia-esto-en-produccion")
ALGORITHM    = "HS256"
TOKEN_EXPIRE = 480

engine       = create_engine(DATABASE_URL)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base         = declarative_base()
Base.metadata.create_all(bind=engine)

def get_db():
    db = SessionLocal()
    try:   yield db
    finally: db.close()

# ── MODELS ──────────────────────────────────────────────────
class Sucursal(Base):
    __tablename__ = "sucursales"
    id         = Column(Integer, primary_key=True)
    nombre     = Column(String(100), nullable=False)
    direccion  = Column(String(200))
    activa     = Column(Boolean, default=True)
    creado_en  = Column(DateTime, default=datetime.now)
    usuarios   = relationship("Usuario", back_populates="sucursal")
    registros  = relationship("Registro", back_populates="sucursal")
    reg_banco  = relationship("RegistroBanco", back_populates="sucursal")

class Usuario(Base):
    __tablename__ = "usuarios"
    id            = Column(Integer, primary_key=True)
    nombre        = Column(String(150), nullable=False)
    email         = Column(String(150), unique=True, nullable=False)
    password_hash = Column(String(255), nullable=False)
    rol           = Column(String(20), nullable=False)
    sucursal_id   = Column(Integer, ForeignKey("sucursales.id"), nullable=True)
    activo        = Column(Boolean, default=True)
    creado_en     = Column(DateTime, default=datetime.now)
    sucursal      = relationship("Sucursal", back_populates="usuarios")
    registros     = relationship("Registro", back_populates="usuario")
    reg_banco     = relationship("RegistroBanco", back_populates="usuario")

class Servicio(Base):
    __tablename__ = "servicios"
    id        = Column(Integer, primary_key=True)
    nombre    = Column(String(100), nullable=False)
    activo    = Column(Boolean, default=True)
    registros = relationship("Registro", back_populates="servicio")

class Registro(Base):
    __tablename__ = "registros"
    id          = Column(Integer, primary_key=True)
    usuario_id  = Column(Integer, ForeignKey("usuarios.id"), nullable=False)
    sucursal_id = Column(Integer, ForeignKey("sucursales.id"), nullable=False)
    servicio_id = Column(Integer, ForeignKey("servicios.id"), nullable=False)
    valor       = Column(Numeric(10, 2), nullable=False)
    fecha       = Column(Date, nullable=False, default=date.today)
    hora        = Column(Time, nullable=False, default=lambda: datetime.now().time())
    creado_en   = Column(DateTime, default=datetime.now)
    usuario     = relationship("Usuario", back_populates="registros")
    sucursal    = relationship("Sucursal", back_populates="registros")
    servicio    = relationship("Servicio", back_populates="registros")

class RegistroBanco(Base):
    __tablename__ = "registros_banco"
    id              = Column(Integer, primary_key=True)
    usuario_id      = Column(Integer, ForeignKey("usuarios.id"), nullable=False)
    sucursal_id     = Column(Integer, ForeignKey("sucursales.id"), nullable=False)
    banco           = Column(String(20), nullable=False)   # pichincha | guayaquil
    tipo_movimiento = Column(String(20), nullable=False)   # pago_servicio | deposito | retiro
    valor           = Column(Numeric(10, 2), nullable=False)
    descripcion     = Column(String(200))
    fecha           = Column(Date, nullable=False, default=date.today)
    hora            = Column(Time, nullable=False, default=lambda: datetime.now().time())
    creado_en       = Column(DateTime, default=datetime.now)
    usuario         = relationship("Usuario", back_populates="reg_banco")
    sucursal        = relationship("Sucursal", back_populates="reg_banco")

# ── SEGURIDAD ────────────────────────────────────────────────
pwd_context   = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/auth/login")

def verify_password(plain, hashed): return pwd_context.verify(plain, hashed)
def hash_password(p): return pwd_context.hash(p)
def create_token(data, exp=None):
    d = data.copy()
    d["exp"] = datetime.utcnow() + (exp or timedelta(minutes=15))
    return jwt.encode(d, SECRET_KEY, algorithm=ALGORITHM)

def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        uid = int(payload.get("sub"))
    except:
        raise HTTPException(status_code=401, detail="Token inválido")
    user = db.query(Usuario).filter(Usuario.id == uid, Usuario.activo == True).first()
    if not user: raise HTTPException(status_code=401, detail="Usuario no encontrado")
    return user

def require_admin(u: Usuario = Depends(get_current_user)):
    if u.rol != "admin": raise HTTPException(status_code=403, detail="Solo admin")
    return u

# ── SCHEMAS ──────────────────────────────────────────────────
class SucursalCreate(BaseModel):
    nombre: str
    direccion: Optional[str] = None

class SucursalOut(BaseModel):
    id: int; nombre: str; direccion: Optional[str]; activa: bool
    class Config: from_attributes = True

class UsuarioCreate(BaseModel):
    nombre: str; email: str; password: str
    rol: str; sucursal_id: Optional[int] = None

class UsuarioOut(BaseModel):
    id: int; nombre: str; email: str; rol: str
    sucursal_id: Optional[int]; activo: bool
    class Config: from_attributes = True

class ServicioOut(BaseModel):
    id: int; nombre: str
    class Config: from_attributes = True

class RegistroCreate(BaseModel):
    servicio_id: int; valor: float

class RegistroBancoCreate(BaseModel):
    banco: str           # pichincha | guayaquil
    tipo_movimiento: str # pago_servicio | deposito | retiro
    valor: float
    descripcion: Optional[str] = None

# ── APP ──────────────────────────────────────────────────────
app = FastAPI(title="Registro Diario v2", version="2.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"],
                   allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

# ── AUTH ─────────────────────────────────────────────────────
@app.post("/auth/login")
def login(form: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(Usuario).filter(Usuario.email == form.username, Usuario.activo == True).first()
    if not user or not verify_password(form.password, user.password_hash):
        raise HTTPException(status_code=400, detail="Credenciales incorrectas")
    token = create_token({"sub": str(user.id)}, timedelta(minutes=TOKEN_EXPIRE))
    return {
        "access_token": token, "token_type": "bearer",
        "usuario": {
            "id": user.id, "nombre": user.nombre, "email": user.email,
            "rol": user.rol, "sucursal_id": user.sucursal_id,
            "sucursal_nombre": user.sucursal.nombre if user.sucursal else None
        }
    }

@app.get("/auth/me")
def me(u: Usuario = Depends(get_current_user)):
    return {"id": u.id, "nombre": u.nombre, "email": u.email, "rol": u.rol,
            "sucursal_id": u.sucursal_id,
            "sucursal_nombre": u.sucursal.nombre if u.sucursal else None}

# ── SUCURSALES ───────────────────────────────────────────────
@app.get("/sucursales", response_model=List[SucursalOut])
def listar_sucursales(db: Session = Depends(get_db), _=Depends(get_current_user)):
    return db.query(Sucursal).filter(Sucursal.activa == True).all()

@app.post("/sucursales", response_model=SucursalOut)
def crear_sucursal(data: SucursalCreate, db: Session = Depends(get_db), _=Depends(require_admin)):
    s = Sucursal(**data.dict()); db.add(s); db.commit(); db.refresh(s); return s

@app.delete("/sucursales/{id}")
def del_sucursal(id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    s = db.query(Sucursal).filter(Sucursal.id == id).first()
    if not s: raise HTTPException(404)
    s.activa = False; db.commit(); return {"ok": True}

# ── USUARIOS ─────────────────────────────────────────────────
@app.get("/usuarios", response_model=List[UsuarioOut])
def listar_usuarios(db: Session = Depends(get_db), _=Depends(require_admin)):
    return db.query(Usuario).filter(Usuario.activo == True).all()

@app.post("/usuarios", response_model=UsuarioOut)
def crear_usuario(data: UsuarioCreate, db: Session = Depends(get_db), _=Depends(require_admin)):
    if db.query(Usuario).filter(Usuario.email == data.email).first():
        raise HTTPException(400, "Email ya registrado")
    u = Usuario(nombre=data.nombre, email=data.email,
                password_hash=hash_password(data.password),
                rol=data.rol, sucursal_id=data.sucursal_id)
    db.add(u); db.commit(); db.refresh(u); return u

@app.delete("/usuarios/{id}")
def del_usuario(id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    u = db.query(Usuario).filter(Usuario.id == id).first()
    if not u: raise HTTPException(404)
    u.activo = False; db.commit(); return {"ok": True}

# ── SERVICIOS ────────────────────────────────────────────────
@app.get("/servicios", response_model=List[ServicioOut])
def listar_servicios(db: Session = Depends(get_db), _=Depends(get_current_user)):
    return db.query(Servicio).filter(Servicio.activo == True).all()

# ── REGISTROS DIARIOS ────────────────────────────────────────
@app.post("/registros")
def crear_registro(data: RegistroCreate, db: Session = Depends(get_db),
                   u: Usuario = Depends(get_current_user)):
    if not u.sucursal_id: raise HTTPException(400, "Sin sucursal asignada")
    r = Registro(usuario_id=u.id, sucursal_id=u.sucursal_id,
                 servicio_id=data.servicio_id, valor=data.valor, fecha=date.today())
    db.add(r); db.commit(); db.refresh(r)
    return {"id": r.id, "servicio": r.servicio.nombre,
            "valor": float(r.valor), "hora": str(r.hora)}

@app.get("/registros/hoy")
def registros_hoy(db: Session = Depends(get_db), u: Usuario = Depends(get_current_user)):
    q = db.query(Registro).filter(Registro.fecha == date.today())
    if u.rol != "admin": q = q.filter(Registro.sucursal_id == u.sucursal_id)
    rows = q.all()
    total = sum(float(r.valor) for r in rows)
    return {
        "registros": [{"id": r.id, "servicio": r.servicio.nombre,
                        "valor": float(r.valor), "hora": str(r.hora)} for r in rows],
        "total": total, "fecha": str(date.today())
    }

@app.get("/registros/fecha")
def registros_fecha(fecha: str, db: Session = Depends(get_db),
                    u: Usuario = Depends(get_current_user)):
    target = date.fromisoformat(fecha)
    q = db.query(Registro).filter(Registro.fecha == target)
    if u.rol != "admin": q = q.filter(Registro.sucursal_id == u.sucursal_id)
    rows = q.all()
    total = sum(float(r.valor) for r in rows)
    return {
        "registros": [{"id": r.id, "servicio": r.servicio.nombre,
                        "valor": float(r.valor), "hora": str(r.hora)} for r in rows],
        "total": total, "fecha": fecha
    }

# ── REGISTROS BANCO ──────────────────────────────────────────
@app.post("/banco/registro")
def crear_registro_banco(data: RegistroBancoCreate, db: Session = Depends(get_db),
                         u: Usuario = Depends(get_current_user)):
    if not u.sucursal_id: raise HTTPException(400, "Sin sucursal asignada")
    if data.banco not in ("pichincha", "guayaquil"):
        raise HTTPException(400, "Banco inválido")
    if data.tipo_movimiento not in ("pago_servicio", "deposito", "retiro"):
        raise HTTPException(400, "Tipo de movimiento inválido")
    r = RegistroBanco(
        usuario_id=u.id, sucursal_id=u.sucursal_id,
        banco=data.banco, tipo_movimiento=data.tipo_movimiento,
        valor=data.valor, descripcion=data.descripcion, fecha=date.today()
    )
    db.add(r); db.commit(); db.refresh(r)
    return {"id": r.id, "banco": r.banco, "tipo_movimiento": r.tipo_movimiento,
            "valor": float(r.valor), "hora": str(r.hora)}

@app.get("/banco/hoy/{banco}")
def banco_hoy(banco: str, db: Session = Depends(get_db), u: Usuario = Depends(get_current_user)):
    if banco not in ("pichincha", "guayaquil"): raise HTTPException(400, "Banco inválido")
    q = db.query(RegistroBanco).filter(
        RegistroBanco.fecha == date.today(),
        RegistroBanco.banco == banco
    )
    if u.rol != "admin": q = q.filter(RegistroBanco.sucursal_id == u.sucursal_id)
    rows = q.all()
    ingresos = [r for r in rows if r.tipo_movimiento in ("pago_servicio", "deposito")]
    retiros  = [r for r in rows if r.tipo_movimiento == "retiro"]
    return {
        "banco": banco,
        "fecha": str(date.today()),
        "ingresos": [{"id": r.id, "tipo": r.tipo_movimiento, "valor": float(r.valor),
                       "descripcion": r.descripcion, "hora": str(r.hora)} for r in ingresos],
        "retiros":  [{"id": r.id, "tipo": r.tipo_movimiento, "valor": float(r.valor),
                       "descripcion": r.descripcion, "hora": str(r.hora)} for r in retiros],
        "total_ingresos": sum(float(r.valor) for r in ingresos),
        "total_retiros":  sum(float(r.valor) for r in retiros),
    }

@app.get("/banco/fecha/{banco}")
def banco_fecha(banco: str, fecha: str, db: Session = Depends(get_db),
                u: Usuario = Depends(get_current_user)):
    if banco not in ("pichincha", "guayaquil"): raise HTTPException(400)
    target = date.fromisoformat(fecha)
    q = db.query(RegistroBanco).filter(
        RegistroBanco.fecha == target, RegistroBanco.banco == banco
    )
    if u.rol != "admin": q = q.filter(RegistroBanco.sucursal_id == u.sucursal_id)
    rows = q.all()
    ingresos = [r for r in rows if r.tipo_movimiento in ("pago_servicio", "deposito")]
    retiros  = [r for r in rows if r.tipo_movimiento == "retiro"]
    return {
        "ingresos": [{"id": r.id, "tipo": r.tipo_movimiento, "valor": float(r.valor),
                       "descripcion": r.descripcion, "hora": str(r.hora)} for r in ingresos],
        "retiros":  [{"id": r.id, "tipo": r.tipo_movimiento, "valor": float(r.valor),
                       "descripcion": r.descripcion, "hora": str(r.hora)} for r in retiros],
        "total_ingresos": sum(float(r.valor) for r in ingresos),
        "total_retiros":  sum(float(r.valor) for r in retiros),
    }

# ── CAJA USUARIO (resumen del día) ───────────────────────────
@app.get("/caja/resumen")
def caja_resumen(fecha: Optional[str] = None, db: Session = Depends(get_db),
                 u: Usuario = Depends(get_current_user)):
    target = date.fromisoformat(fecha) if fecha else date.today()
    # Servicios
    qr = db.query(Registro).filter(Registro.fecha == target)
    if u.rol != "admin": qr = qr.filter(Registro.sucursal_id == u.sucursal_id)
    total_servicios = sum(float(r.valor) for r in qr.all())
    # Bancos
    def banco_totales(banco):
        qb = db.query(RegistroBanco).filter(
            RegistroBanco.fecha == target, RegistroBanco.banco == banco)
        if u.rol != "admin": qb = qb.filter(RegistroBanco.sucursal_id == u.sucursal_id)
        rows = qb.all()
        ing = sum(float(r.valor) for r in rows if r.tipo_movimiento in ("pago_servicio","deposito"))
        ret = sum(float(r.valor) for r in rows if r.tipo_movimiento == "retiro")
        return ing, ret
    pi, pr = banco_totales("pichincha")
    gi, gr = banco_totales("guayaquil")
    return {
        "fecha": str(target),
        "total_servicios": total_servicios,
        "pichincha": {"ingresos": pi, "retiros": pr},
        "guayaquil": {"ingresos": gi, "retiros": gr},
        "gran_total": total_servicios + (pi - pr) + (gi - gr)
    }

# ── REPORTES ADMIN ───────────────────────────────────────────
@app.get("/reportes/diario")
def reporte_diario(fecha: Optional[str] = None, db: Session = Depends(get_db),
                   _=Depends(require_admin)):
    target = date.fromisoformat(fecha) if fecha else date.today()
    rows = db.execute(text("""
        SELECT s.id, s.nombre,
               COUNT(r.id) AS total_registros,
               COALESCE(SUM(r.valor), 0) AS total_valor
        FROM sucursales s
        LEFT JOIN registros r ON r.sucursal_id = s.id AND r.fecha = :fecha
        WHERE s.activa = TRUE
        GROUP BY s.id, s.nombre ORDER BY s.nombre
    """), {"fecha": target}).fetchall()
    sucursales = [{"sucursal_id": r[0], "sucursal": r[1],
                    "total_registros": r[2], "total_valor": float(r[3])} for r in rows]
    return {"fecha": str(target), "sucursales": sucursales,
            "gran_total": sum(r["total_valor"] for r in sucursales)}

@app.get("/reportes/banco")
def reporte_banco(banco: str, fecha_inicio: Optional[str] = None,
                  fecha_fin: Optional[str] = None, sucursal_id: Optional[int] = None,
                  usuario_id: Optional[int] = None, db: Session = Depends(get_db),
                  _=Depends(require_admin)):
    fi = date.fromisoformat(fecha_inicio) if fecha_inicio else date.today()
    ff = date.fromisoformat(fecha_fin) if fecha_fin else date.today()
    filters = "WHERE rb.banco = :banco AND rb.fecha BETWEEN :fi AND :ff AND s.activa = TRUE"
    params  = {"banco": banco, "fi": fi, "ff": ff}
    if sucursal_id:
        filters += " AND rb.sucursal_id = :sid"; params["sid"] = sucursal_id
    if usuario_id:
        filters += " AND rb.usuario_id = :uid"; params["uid"] = usuario_id
    rows = db.execute(text(f"""
        SELECT rb.fecha, s.nombre AS sucursal, u.nombre AS usuario,
               rb.tipo_movimiento, SUM(rb.valor) AS total
        FROM registros_banco rb
        JOIN sucursales s ON s.id = rb.sucursal_id
        JOIN usuarios u ON u.id = rb.usuario_id
        {filters}
        GROUP BY rb.fecha, s.nombre, u.nombre, rb.tipo_movimiento
        ORDER BY rb.fecha DESC, s.nombre
    """), params).fetchall()
    resultado = [{"fecha": str(r[0]), "sucursal": r[1], "usuario": r[2],
                   "tipo": r[3], "total": float(r[4])} for r in rows]
    # Totales
    ingresos = sum(r["total"] for r in resultado if r["tipo"] in ("pago_servicio","deposito"))
    retiros  = sum(r["total"] for r in resultado if r["tipo"] == "retiro")
    return {"banco": banco, "registros": resultado,
            "total_ingresos": ingresos, "total_retiros": retiros}

@app.get("/reportes/semanal")
def reporte_semanal(db: Session = Depends(get_db), _=Depends(require_admin)):
    rows = db.execute(text("""
        SELECT DATE_TRUNC('week', r.fecha)::date AS semana,
               s.nombre, COALESCE(SUM(r.valor), 0) AS total
        FROM registros r JOIN sucursales s ON s.id = r.sucursal_id
        WHERE r.fecha >= CURRENT_DATE - INTERVAL '4 weeks'
        GROUP BY DATE_TRUNC('week', r.fecha), s.nombre
        ORDER BY semana DESC, s.nombre
    """)).fetchall()
    return [{"semana": str(r[0]), "sucursal": r[1], "total": float(r[2])} for r in rows]

@app.get("/reportes/mensual")
def reporte_mensual(db: Session = Depends(get_db), _=Depends(require_admin)):
    rows = db.execute(text("""
        SELECT DATE_TRUNC('month', r.fecha)::date AS mes,
               s.nombre, COALESCE(SUM(r.valor), 0) AS total
        FROM registros r JOIN sucursales s ON s.id = r.sucursal_id
        WHERE r.fecha >= CURRENT_DATE - INTERVAL '12 months'
        GROUP BY DATE_TRUNC('month', r.fecha), s.nombre
        ORDER BY mes DESC, s.nombre
    """)).fetchall()
    return [{"mes": str(r[0]), "sucursal": r[1], "total": float(r[2])} for r in rows]

# ── REPORTE EXCEL GENERAL ────────────────────────────────────
@app.get("/reportes/excel")
def reporte_excel(fecha_inicio: str, fecha_fin: str,
                  sucursal_id: Optional[int] = None,
                  db: Session = Depends(get_db), _=Depends(require_admin)):
    fi = date.fromisoformat(fecha_inicio)
    ff = date.fromisoformat(fecha_fin)

    # Estilos
    hdr_fill   = PatternFill("solid", fgColor="1E3058")
    hdr_font   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    sub_fill   = PatternFill("solid", fgColor="2563EB")
    sub_font   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    total_fill = PatternFill("solid", fgColor="0F1726")
    total_font = Font(bold=True, color="60A5FA", name="Arial", size=10)
    body_font  = Font(name="Arial", size=10)
    alt_fill   = PatternFill("solid", fgColor="F0F4FF")
    thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    center = Alignment(horizontal='center', vertical='center')

    def set_header(ws, title, cols):
        ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
        c = ws["A1"]
        c.value = title
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=13)
        c.fill = PatternFill("solid", fgColor="0A0F1E")
        c.alignment = center
        ws.row_dimensions[1].height = 28
        for i, col in enumerate(cols, 1):
            cell = ws.cell(row=2, column=i, value=col)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = center; cell.border = thin
        ws.row_dimensions[2].height = 20

    def write_rows(ws, data, start=3):
        for ri, row in enumerate(data):
            fill = alt_fill if ri % 2 == 0 else PatternFill()
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=start+ri, column=ci, value=val)
                cell.font = body_font; cell.border = thin
                if fill.fill_type: cell.fill = fill
        return start + len(data)

    def write_total_row(ws, row_idx, cols_count, label, value):
        ws.merge_cells(f"A{row_idx}:{get_column_letter(cols_count-1)}{row_idx}")
        lbl = ws.cell(row=row_idx, column=1, value=label)
        lbl.font = total_font; lbl.fill = total_fill
        lbl.alignment = Alignment(horizontal='right'); lbl.border = thin
        val = ws.cell(row=row_idx, column=cols_count, value=value)
        val.font = Font(bold=True, color="10B981", name="Arial", size=10)
        val.fill = total_fill; val.border = thin
        val.number_format = '"$"#,##0.00'

    wb = openpyxl.Workbook()

    # ── HOJA 1: SERVICIOS ─────────────────────────────────────
    ws1 = wb.active; ws1.title = "Servicios"
    suc_filter = "AND r.sucursal_id = :sid" if sucursal_id else ""
    params1 = {"fi": fi, "ff": ff}
    if sucursal_id: params1["sid"] = sucursal_id
    rows1 = db.execute(text(f"""
        SELECT r.fecha, s.nombre AS sucursal, u.nombre AS usuario,
               sv.nombre AS servicio, r.valor, r.hora
        FROM registros r
        JOIN sucursales s ON s.id = r.sucursal_id
        JOIN usuarios u ON u.id = r.usuario_id
        JOIN servicios sv ON sv.id = r.servicio_id
        WHERE r.fecha BETWEEN :fi AND :ff {suc_filter}
        ORDER BY r.fecha DESC, s.nombre, r.hora
    """), params1).fetchall()
    cols1 = ["Fecha", "Sucursal", "Usuario", "Servicio", "Valor ($)", "Hora"]
    set_header(ws1, f"Reporte de Servicios — {fecha_inicio} al {fecha_fin}", cols1)
    data1 = [(str(r[0]), r[1], r[2], r[3], float(r[4]), str(r[5])[:5]) for r in rows1]
    last1 = write_rows(ws1, data1)
    total1 = sum(float(r[4]) for r in rows1)
    write_total_row(ws1, last1, len(cols1), "TOTAL SERVICIOS", total1)
    ws1.column_dimensions["A"].width = 12; ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 25; ws1.column_dimensions["D"].width = 30
    ws1.column_dimensions["E"].width = 14; ws1.column_dimensions["F"].width = 10
    for row in ws1.iter_rows(min_row=3):
        if row[4].value and isinstance(row[4].value, float):
            row[4].number_format = '"$"#,##0.00'

    # ── HOJA 2: BANCO PICHINCHA ───────────────────────────────
    ws2 = wb.create_sheet("Banco Pichincha")
    suc_filter2 = "AND rb.sucursal_id = :sid" if sucursal_id else ""
    params2 = {"fi": fi, "ff": ff, "banco": "pichincha"}
    if sucursal_id: params2["sid"] = sucursal_id
    rows2 = db.execute(text(f"""
        SELECT rb.fecha, s.nombre, u.nombre,
               rb.tipo_movimiento, rb.valor, rb.descripcion, rb.hora
        FROM registros_banco rb
        JOIN sucursales s ON s.id = rb.sucursal_id
        JOIN usuarios u ON u.id = rb.usuario_id
        WHERE rb.banco = :banco AND rb.fecha BETWEEN :fi AND :ff {suc_filter2}
        ORDER BY rb.fecha DESC, s.nombre, rb.hora
    """), params2).fetchall()
    TIPO_LABELS = {"pago_servicio": "Pago de servicio", "deposito": "Depósito", "retiro": "Retiro"}
    cols2 = ["Fecha", "Sucursal", "Usuario", "Tipo", "Valor ($)", "Descripción", "Hora"]
    set_header(ws2, f"Banco Pichincha — {fecha_inicio} al {fecha_fin}", cols2)
    data2 = [(str(r[0]), r[1], r[2], TIPO_LABELS.get(r[3], r[3]),
               float(r[4]), r[5] or "", str(r[6])[:5]) for r in rows2]
    last2 = write_rows(ws2, data2)
    ing2 = sum(float(r[4]) for r in rows2 if r[3] in ("pago_servicio","deposito"))
    ret2 = sum(float(r[4]) for r in rows2 if r[3] == "retiro")
    write_total_row(ws2, last2,   len(cols2), "TOTAL INGRESOS (Depósitos + Pagos)", ing2)
    write_total_row(ws2, last2+1, len(cols2), "TOTAL RETIROS", ret2)
    for w, col in zip([12,20,25,18,14,25,10], ["A","B","C","D","E","F","G"]):
        ws2.column_dimensions[col].width = w
    for row in ws2.iter_rows(min_row=3):
        if row[4].value and isinstance(row[4].value, float):
            row[4].number_format = '"$"#,##0.00'

    # ── HOJA 3: BANCO GUAYAQUIL ───────────────────────────────
    ws3 = wb.create_sheet("Banco Guayaquil")
    params3 = {**params2, "banco": "guayaquil"}
    rows3 = db.execute(text(f"""
        SELECT rb.fecha, s.nombre, u.nombre,
               rb.tipo_movimiento, rb.valor, rb.descripcion, rb.hora
        FROM registros_banco rb
        JOIN sucursales s ON s.id = rb.sucursal_id
        JOIN usuarios u ON u.id = rb.usuario_id
        WHERE rb.banco = :banco AND rb.fecha BETWEEN :fi AND :ff {suc_filter2}
        ORDER BY rb.fecha DESC, s.nombre, rb.hora
    """), params3).fetchall()
    set_header(ws3, f"Banco Guayaquil — {fecha_inicio} al {fecha_fin}", cols2)
    data3 = [(str(r[0]), r[1], r[2], TIPO_LABELS.get(r[3], r[3]),
               float(r[4]), r[5] or "", str(r[6])[:5]) for r in rows3]
    last3 = write_rows(ws3, data3)
    ing3 = sum(float(r[4]) for r in rows3 if r[3] in ("pago_servicio","deposito"))
    ret3 = sum(float(r[4]) for r in rows3 if r[3] == "retiro")
    write_total_row(ws3, last3,   len(cols2), "TOTAL INGRESOS (Depósitos + Pagos)", ing3)
    write_total_row(ws3, last3+1, len(cols2), "TOTAL RETIROS", ret3)
    for w, col in zip([12,20,25,18,14,25,10], ["A","B","C","D","E","F","G"]):
        ws3.column_dimensions[col].width = w
    for row in ws3.iter_rows(min_row=3):
        if row[4].value and isinstance(row[4].value, float):
            row[4].number_format = '"$"#,##0.00'

    # ── HOJA 4: RESUMEN GENERAL ───────────────────────────────
    ws4 = wb.create_sheet("Resumen General")
    ws4.merge_cells("A1:D1")
    c = ws4["A1"]; c.value = f"Resumen General — {fecha_inicio} al {fecha_fin}"
    c.font = Font(bold=True, color="FFFFFF", name="Arial", size=14)
    c.fill = PatternFill("solid", fgColor="0A0F1E"); c.alignment = center
    ws4.row_dimensions[1].height = 30
    total_pichincha = ing2 - ret2
    total_guayaquil = ing3 - ret3
    gran_total = total1 + total_pichincha + total_guayaquil
    resumen = [
        ("Total Caja de Servicios", total1),
        ("Total Pichincha (Ingresos - Retiros)", total_pichincha),
        ("Total Guayaquil (Ingresos - Retiros)", total_guayaquil),
        ("GRAN TOTAL GENERAL", gran_total),
    ]
    for i, (label, value) in enumerate(resumen):
        r_idx = i + 2
        lbl = ws4.cell(row=r_idx, column=1, value=label)
        lbl.font = Font(bold=(i==len(resumen)-1), name="Arial", size=11)
        lbl.fill = PatternFill("solid", fgColor="1E3058") if i == len(resumen)-1 else PatternFill()
        lbl.border = thin
        val = ws4.cell(row=r_idx, column=2, value=value)
        val.font = Font(bold=(i==len(resumen)-1),color="10B981" if value >= 0 else "EF4444",name="Arial",size=11)
        val.number_format = '"$"#,##0.00'; val.border = thin
        ws4.row_dimensions[r_idx].height = 22
    ws4.column_dimensions["A"].width = 42; ws4.column_dimensions["B"].width = 18

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    filename = f"reporte_{fecha_inicio}_{fecha_fin}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": f'attachment; filename="{filename}"'})
